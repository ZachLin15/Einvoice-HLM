import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import warnings

# Suppress all Excel-related warnings for automated processing
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', message='.*Data Validation.*')
warnings.filterwarnings('ignore', message='.*Workbook contains no default style.*')


def copy_data_between_files(source_file, submission_file, output_file):
    """
    Copy data from source file sheets to submission file sheets and save as new file.

    Args:
        source_file: Path to exportM file
        submission_file: Path to submission file
        output_file: Path to save the new combined file
    """

    # Sheets to copy data between
    sheets_to_copy = [
        'Documents',
        'DocumentLineItems',
        'LineItemsTaxes',
        'DocumentTotalTax'
    ]

    try:
        print("Loading files...")
        print(f"Source file: {source_file}")
        print(f"Submission file: {submission_file}")

        # Load both workbooks
        source_wb = load_workbook(source_file)
        submission_wb = load_workbook(submission_file)

        print(f"Source sheets available: {source_wb.sheetnames}")
        print(f"Submission sheets available: {submission_wb.sheetnames}")

        # Process each sheet
        for sheet_name in sheets_to_copy:
            if sheet_name in source_wb.sheetnames and sheet_name in submission_wb.sheetnames:
                print(f"\nCopying data for sheet: {sheet_name}")

                # Get sheets
                source_sheet = source_wb[sheet_name]
                submission_sheet = submission_wb[sheet_name]

                # Find data range in source (skip header row)
                max_row = source_sheet.max_row
                max_col = source_sheet.max_column

                print(f"Source data: {max_row} rows x {max_col} columns")

                # Copy data from source to submission (skip source header, start at row 6 in submission)
                data_rows_copied = 0
                for row in range(2, max_row + 1):  # Skip header row in source (start from row 2)
                    for col in range(1, max_col + 1):
                        source_cell = source_sheet.cell(row=row, column=col)
                        # Paste starting at row 6 in submission (keeping top 5 rows)
                        submission_cell = submission_sheet.cell(row=row + 4, column=col)
                        submission_cell.value = source_cell.value
                    data_rows_copied += 1

                print(f"Copied {data_rows_copied} rows of data to {sheet_name}")

            elif sheet_name not in source_wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in source file")
            elif sheet_name not in submission_wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in submission file")

        # Save as new Excel file
        print(f"\nSaving combined file as: {output_file}")
        submission_wb.save(output_file)

        print("✓ Data copy completed successfully!")
        print(f"✓ New file saved: {output_file}")
        print("✓ Original files remain unchanged")

    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
    except Exception as e:
        print(f"Error occurred: {e}")


def copy_data_pandas_method(source_file, submission_file, output_file):
    """
    Alternative method using pandas for simpler data copying
    """

    sheets_to_copy = [
        'Documents',
        'DocumentLineItems',
        'LineItemsTaxes',
        'DocumentTotalTax'
    ]

    try:
        print("Loading files with pandas...")

        # Read data from both files
        source_data = pd.read_excel(source_file, sheet_name=None)
        submission_data = pd.read_excel(submission_file, sheet_name=None)

        print(f"Source sheets: {list(source_data.keys())}")
        print(f"Submission sheets: {list(submission_data.keys())}")

        # Create output data dictionary starting with submission structure
        output_data = submission_data.copy()

        # Process each sheet
        for sheet_name in sheets_to_copy:
            if sheet_name in source_data and sheet_name in submission_data:
                print(f"\nProcessing sheet: {sheet_name}")

                # Get source data (skip header row)
                source_df = source_data[sheet_name]
                submission_df = submission_data[sheet_name]

                if len(source_df) > 1:  # Has data beyond header
                    # Keep top 5 rows from submission file
                    submission_top5 = submission_df.head(5)

                    # Get data from source (skip first row which is header)
                    source_data_only = source_df.iloc[1:].reset_index(drop=True)

                    # Combine: submission top 5 rows + source data
                    combined_df = pd.concat([submission_top5, source_data_only], ignore_index=True)
                    output_data[sheet_name] = combined_df

                    print(f"Combined: 5 header rows + {len(source_data_only)} data rows")
                else:
                    print(f"No data to copy from source for sheet: {sheet_name}")
            else:
                if sheet_name not in source_data:
                    print(f"Warning: Sheet '{sheet_name}' not found in source file")
                if sheet_name not in submission_data:
                    print(f"Warning: Sheet '{sheet_name}' not found in submission file")

        # Save to new file
        print(f"\nSaving to: {output_file}")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in output_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("✓ Pandas method completed successfully!")
        print(f"✓ New file saved: {output_file}")
        print("✓ Original files remain unchanged")

    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
    except Exception as e:
        print(f"Error occurred: {e}")


if __name__ == "__main__":
    # File paths - update these with your actual file paths
    source_file = r"C:\Users\Admin\Documents\Work\M-Einvoice\exportM.xlsx"  # Source file
    submission_file = r"C:\Users\Admin\Documents\Work\M-Einvoice\BatchSubmission-v2.xlsx"  # Submission file
    output_file = r"C:\Users\Admin\Documents\Work\M-Einvoice\BatchSubmission-Combined.xlsx"  # New output file

    # Check if files exist
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' not found!")
        exit(1)

    if not os.path.exists(submission_file):
        print(f"Error: Submission file '{submission_file}' not found!")
        exit(1)


    print("1. Copy with openpyxl (preserves formatting)")





    copy_data_between_files(source_file, submission_file, output_file)
